#!/usr/bin/env python
# coding: utf-8

# Transfer of Biefselect results from PDF to excel file

# Import packages
import os
import pdftotext
import pandas as pd
import re
import datetime
import numpy as np
import shutil
import sys
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
if sys.version_info[0] < 3: 
    from StringIO import StringIO
else:
    from io import StringIO

# Packages for dialog boxes
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog

# Dictionary with code value mappings:
Vetbedekking = {'1-':0.33, 
                '10': 1.00, 
                '1+': 1.33, 
                '2-': 1.66, 
                '20': 2.00, 
                '2+': 2.33, 
                '3-': 2.66, 
                '30': 3.00, 
                '3+': 3.33}

Vleesbedekking = {'S': 5.66,
                  'E+': 5.33, 
                  'E0': 5.00, 
                  'E-': 4.66, 
                  'U+': 4.33, 
                  'U0': 4.00, 
                  'U-': 3.66, 
                  'R+': 3.33, 
                  'R0': 3.00, 
                  'R-': 2.66}

# Constants
FactorCold = 0.98
Intercept = 53.662
SlopeCold = 0.01523
SlopeMeat = 1.255
SlopeFat = -1.202

# Targetfile
gewichtexcel = "Gewicht2.xlsx"

# Define the columns and their respective formats
columns_formats = {
    0: ('d-mm-yy', None), # Column B: date, dd-mm-yy
    1: (numbers.FORMAT_NUMBER, None), # Column C: number, 0 digits behind decimal point
    2: ('0.0', None), # Column D, I, K, O: number, 1 digit behind decimal point
    3: (numbers.FORMAT_TEXT, None), # Column E, G: text
    4: (numbers.FORMAT_NUMBER_00, None), # Column F, H, J: number, 2 digits behind decimal point
    5: (numbers.FORMAT_TEXT, None),
    6: (numbers.FORMAT_NUMBER_00, None),
    7: ('0.0', None),
    8: (numbers.FORMAT_NUMBER_00, None),
    9: ('0.0', None),
    13: ('0.0', None),
}

# Functions

# Find default download folder
def get_download_path():
    """Returns the default downloads path for linux or windows"""
    if os.name == 'nt':
        import winreg
        sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
        downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
            location = winreg.QueryValueEx(key, downloads_guid)[0]
        return location
    else:
        return os.path.join(os.path.expanduser('~'), 'downloads')
    
# Find most recent PDF    
def find_most_recent_pdf(folder_path):
    pdf_files = [f for f in os.listdir(folder_path) if f.endswith('.pdf')]

    if not pdf_files:
        return None

    # Get the full paths of the PDF files
    pdf_paths = [os.path.join(folder_path, pdf_file) for pdf_file in pdf_files]

    # Find the most recently modified PDF file
    most_recent_pdf = max(pdf_paths, key=os.path.getmtime)

    return most_recent_pdf

# Show dialog box
def show_dialog_box(title, boxtext):
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    answer = messagebox.askyesno(title, boxtext)

    # Convert the answer to "Yes" or "No"
    answer = "Yes" if answer else "No"

    return answer


# 1. Which input file to process?
# Set downloads folder
PDFdir = get_download_path()


# Which is the most recent PDF?
most_recent_pdf_path = find_most_recent_pdf(PDFdir)
SelectedFilename = os.path.basename(most_recent_pdf_path)

# Aks if this is the correct one
boxtext = "Wil je de slachtresultaten van de file '" + SelectedFilename + "' in de downloads folder importeren?"
title = "Import slachtresultaten"
answer = show_dialog_box(title, boxtext)


# If not correct file: let user select file. Else, set as correct file
if answer == "No":
    root = tk.Tk()
    root.withdraw()

    ImportFile = filedialog.askopenfilename()

elif answer == "Yes": 
    ImportFile = most_recent_pdf_path


# Double check file type: should be PDF! 
file_name, file_extension = os.path.splitext(ImportFile)
if file_extension != ".pdf":
    messagebox.showerror('Error: geen PDF', 'Error: dit is geen PDF file. Selecteer de juiste file!')


# 2. Parse PDF

with open(ImportFile, "rb") as f:
    pdf = pdftotext.PDF(f)
f.close()
textPDF = pdf[0]
textaslist = textPDF.split("\r\n")


# Extract date
# Set pattern - Round brackets around number extracts date.
pattern_date = '(\d{2}-\d{2}-\d{2})'
date_str = re.findall(pattern_date, textaslist[3])
# Extract the date as real date - the format must be the format the date is in.
date_object_pdf = datetime.datetime.strptime(date_str[0], '%d-%m-%y')


# Extract table with data --> to dataframe
# Table always between 'correctie' & 'totalen'
pattern_table = 'correctie([\s\S]*)totalen'
table_match = re.search(pattern_table, textPDF)

if table_match:
    table_unf = table_match.group(1)
    
    # To replace spaces as separator with pipe
    table_unf = re.sub(r'(?<=\S)[ ]+(?=\S)', '|', table_unf)
    table_unf = re.sub(r'(?<=\n)[ ]+', '', table_unf)
   
    table_unf_aslist = table_unf.split("\r\n")

df = pd.DataFrame([x.split('|') for x in table_unf.split('\r\n')])
# Clean up dataframe

# Get rid of columns/rows we don't need
df_cpy = df.copy()
df_cpy.drop(6,axis = 1, inplace=True)
df_cpy = df_cpy[1:-1]

num_columns = df_cpy.shape[1]
headers = ['volg#',	'Land','ID-code', 'keurmerk', 'geboortedatum', 'Vleesbedekking_classificatie', 'Vetbedekking_classificatie', 'netto', 'tarra'] + [f"correctie{i}" for i in range(1, num_columns - 9 + 1)]
df_cpy.columns = headers
 
# Extract remarks
pattern_remarks = 'Opmerkingen([\s\S]*)Geboortelanden'
remarks_match = re.search(pattern_remarks, textPDF)
if remarks_match:
    remarks_unf = remarks_match.group(1)
    # Split in list on newline character
    remarks_unf_aslist = remarks_unf.split("\r\n")

# 3. Add extracted data to existing excel file 
# Drop columns Land, geboortedatum, Keurmerk, tarra
columns_to_drop_excel = ['Land','geboortedatum','keurmerk', 'tarra'] 
df_cpy.drop(columns_to_drop_excel,axis = 1, inplace=True)

# Replace volg# with sequential number 
df_cpy['volg#'] = range(len(df_cpy))
df_cpy['volg#'] = df_cpy['volg#'] + 1

# Reset index
df_cpy.reset_index(drop=True, inplace=True)

# Add date
formatted_date = date_object_pdf.strftime('%d-%m-%y')
df_cpy.insert(1, 'slachtdatum', formatted_date) 

# Move netto column to 3rd position
netto_column = df_cpy.pop('netto') 
df_cpy.insert(3, 'netto', netto_column) 

# Add column wih vleesbedekkingscore after Vleesbedekkingclassificatie
df_cpy.insert(df_cpy.columns.get_loc('Vleesbedekking_classificatie') + 1, 'Vleesbedekking_score', df_cpy['Vleesbedekking_classificatie'].map(Vleesbedekking))

# Add column wih Vetbedekkingscore after vetbedekkingclassificatie
df_cpy.insert(df_cpy.columns.get_loc('Vetbedekking_classificatie') + 1, 'Vetbedekking_score', df_cpy['Vetbedekking_classificatie'].map(Vetbedekking))

# Add empty column for koud gewicht, Aanhoud%, Levend gewicht, gemiddeld slachtgewicht before correctie column
df_cpy.insert(8, 'Koud gewicht', '') 
df_cpy.insert(9, 'Aanhoud %', '') 
df_cpy.insert(10, 'Levend gewicht', '') 
df_cpy.insert(11, 'Gemiddelde', '') 

# Convert weight column to US format depending on decimal separator system settings windows
df_cpy['netto'] = [x.replace(',', '.') for x in df_cpy['netto']]
df_cpy['netto'] = df_cpy['netto'].astype(float)

# add mean in the bottom
df_cpy.Gemiddelde.iloc[-1] = round(df_cpy.netto.mean(),1)

# Save Back up of the workbook first.
BackupFolder = "BackUp"

if not os.path.exists(BackupFolder):
    os.mkdir(BackupFolder) 

currentTS = datetime.datetime.now().strftime("%Y%b%d-%H%M%S")
BUFilenamePath = "./Backup/" + "Gewicht2_" + currentTS + ".xlsx"

shutil.copyfile(gewichtexcel, BUFilenamePath)

# Load the workbook & worksheet
wb = load_workbook(gewichtexcel)
ws = wb['Gewicht']
# Find the last row with data
last_row = ws.max_row
# Determine the starting row for the new data
start_row = last_row + 1

# Write the new data to the Excel sheet by looping over rows.
# Loop over rows
for index, row in df_cpy.iterrows():
    # Loop over columns in row --> make sure to start at 1 to paste in first column
    for col_index, value in enumerate(row, start=1):
        # There are hidden columns; sigh
        if col_index > 11:
            ws.cell(row=start_row + index, column=col_index+3, value=value)
        else:
            ws.cell(row=start_row + index, column=col_index, value=value)
        
# Add the formulas to the 'total' column
for index in range(len(df_cpy)):
    row_number = start_row + index
    formulaI = f'=D{row_number}*{FactorCold}'
    formulaJ = f'={Intercept}+{SlopeCold}*I{row_number}+{SlopeMeat}*F{row_number}+{SlopeFat}*H{row_number}'
    formulaK = f'=I{row_number}*100/J{row_number}'
    ws.cell(row=row_number, column=9, value=formulaI)
    ws.cell(row=row_number, column=10, value=formulaJ)
    ws.cell(row=row_number, column=11, value=formulaK)


# Right align columns B till O:
al = Alignment(horizontal='right')
all_cells = ws['B{}:O{}'.format(start_row,ws.max_row)]


for row in all_cells:
    for col_index, (number_format, text_format) in columns_formats.items():
        cell = row[col_index]
        cell.number_format = number_format
        if text_format:
            cell.number_format = text_format
        cell.alignment = al

# Save the updated Excel file
wb.save(gewichtexcel)

# Move copy of PDF to PDF folder
os.path.basename(ImportFile)
PDFfolder = "Slachtlijsten_PDF"

if not os.path.exists(PDFfolder):
    os.mkdir(PDFfolder) 

file_name_PDFBU = "./Slachtlijsten_PDF/" + os.path.basename(ImportFile)

shutil.move(ImportFile, file_name_PDFBU)